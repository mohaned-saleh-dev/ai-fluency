#!/usr/bin/env python3
"""
Create consolidated rollout schedule Excel from Zendesk-Salesforce Transition Plan.
Shows agent counts per LOB per day for February 2026.
UPDATED: Feb 2, 2026 with new L2 data and date changes.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from collections import defaultdict

# Output file
OUTPUT_FILE = "/Users/mohaned.saleh/Downloads/SF_Rollout_Agent_Schedule_Feb2026.xlsx"

# Define all LOBs from the document (expanded to ~60)
# Format: (LOB Name, Channel, Country)
LOBS = [
    # ========== PARTNER CARE (L1) - 6 LOBs ==========
    ("Partner Care (L1)", "Email", "UAE"),
    ("Partner Care (L1)", "Email", "KSA"),
    ("Partner Care (L1)", "Chat", "UAE"),
    ("Partner Care (L1)", "Chat", "KSA"),
    ("Partner Care (L1)", "Phone", "UAE"),
    ("Partner Care (L1)", "Phone", "KSA"),
    
    # ========== PARTNER ONBOARDING (L1) - 6 LOBs ==========
    ("Partner Onboarding (L1)", "Email", "UAE"),
    ("Partner Onboarding (L1)", "Email", "KSA"),
    ("Partner Onboarding (L1)", "Chat", "UAE"),
    ("Partner Onboarding (L1)", "Chat", "KSA"),
    ("Partner Onboarding (L1)", "Phone", "UAE"),
    ("Partner Onboarding (L1)", "Phone", "KSA"),
    
    # ========== CUSTOMER CARE (L1) (STANDARD) - 6 LOBs ==========
    ("Customer Care (L1) (Standard)", "Email", "UAE"),
    ("Customer Care (L1) (Standard)", "Email", "KSA"),
    ("Customer Care (L1) (Standard)", "Chat", "UAE"),
    ("Customer Care (L1) (Standard)", "Chat", "KSA"),
    ("Customer Care (L1) (Standard)", "Phone", "UAE"),
    ("Customer Care (L1) (Standard)", "Phone", "KSA"),
    
    # ========== CUSTOMER CARE (L1) (SMART) - 6 LOBs ==========
    ("Customer Care (L1) (Smart)", "Email", "UAE"),
    ("Customer Care (L1) (Smart)", "Email", "KSA"),
    ("Customer Care (L1) (Smart)", "Chat", "UAE"),
    ("Customer Care (L1) (Smart)", "Chat", "KSA"),
    ("Customer Care (L1) (Smart)", "Phone", "UAE"),
    ("Customer Care (L1) (Smart)", "Phone", "KSA"),
    
    # ========== PARTNER CARE L2 TEAMS - 8 LOBs ==========
    ("Partner Technical Support L2", "All", "Cross-country"),
    ("Partner Trust & Safety L2", "All", "Cross-country"),
    ("Partner Finance L2", "All", "Cross-country"),
    ("Partner Manager Escalation L2", "All", "Cross-country"),
    
    # ========== CUSTOMER L2 TEAMS - 10 LOBs ==========
    ("Customer Care L2 (SME)", "All", "Cross-country"),
    ("Customer Govt & Exec Escalations L2", "All", "Cross-country"),
    ("Customer Manager Escalations L2", "All", "Cross-country"),
    ("Customer Technical Support L2", "All", "Cross-country"),
    ("Customer Trust & Safety L2", "All", "Cross-country"),
    ("Customer Finance L2", "All", "Cross-country"),
    ("Customer Risk L2", "All", "Cross-country"),
    ("Customer Collections L2", "All", "Cross-country"),
    ("Credit L2 (Placeholder)", "All", "Cross-country"),
    
    # ========== SUPPORT TEAMS - 5 LOBs ==========
    ("Content Team", "All", "Cross-country"),
    ("Training Team", "All", "Cross-country"),
    ("WFM/RTA Team", "All", "Cross-country"),
    ("Quality Team", "All", "Cross-country"),
    ("OpsX Team", "All", "Cross-country"),
]

def get_lob_key(lob, channel, country):
    """Create a unique key for a LOB."""
    return f"{lob} - {channel} - {country}"

def create_rollout_data():
    """Create the rollout data structure based on the UPDATED Notion document tables."""
    
    # Dictionary to hold agent counts: {date: {lob_key: count}}
    rollout_data = defaultdict(lambda: defaultdict(int))
    
    # Helper function to add rollout data for date ranges
    def add_rollout(start_date, end_date, lob, channel, country, agent_count):
        """Add rollout data for a date range. Once rolled out, agents stay on SF."""
        lob_key = get_lob_key(lob, channel, country)
        current = start_date
        # Set the count starting from start_date through end of Feb
        feb_end = datetime(2026, 2, 28)
        while current <= feb_end:
            if current >= start_date:
                # Keep the maximum count seen (for gradual rollouts)
                rollout_data[current][lob_key] = max(rollout_data[current][lob_key], agent_count)
            current += timedelta(days=1)
    
    # ============================================================
    # PARTNER CARE (L1) - UPDATED DATES
    # ============================================================
    
    # 03/02/2026: Email (UAE + KSA) - 6 agents total (split 3/3)
    add_rollout(datetime(2026, 2, 3), datetime(2026, 2, 3), "Partner Care (L1)", "Email", "UAE", 3)
    add_rollout(datetime(2026, 2, 3), datetime(2026, 2, 3), "Partner Care (L1)", "Email", "KSA", 3)
    
    # 02/02/2026 → 04/02/2026: Soft launch
    add_rollout(datetime(2026, 2, 2), datetime(2026, 2, 4), "Partner Care (L1)", "Phone", "KSA", 2)
    add_rollout(datetime(2026, 2, 2), datetime(2026, 2, 4), "Partner Care (L1)", "Chat", "KSA", 2)
    add_rollout(datetime(2026, 2, 2), datetime(2026, 2, 4), "Partner Care (L1)", "Phone", "UAE", 2)
    add_rollout(datetime(2026, 2, 2), datetime(2026, 2, 4), "Partner Care (L1)", "Chat", "UAE", 2)
    
    # 05-06 Feb: Chat (UAE) 20% - 8 agents
    add_rollout(datetime(2026, 2, 5), datetime(2026, 2, 6), "Partner Care (L1)", "Chat", "UAE", 8)
    
    # 07-08 Feb: Chat (UAE) 50% - 15 agents
    add_rollout(datetime(2026, 2, 7), datetime(2026, 2, 8), "Partner Care (L1)", "Chat", "UAE", 15)
    
    # 09-10 Feb: Chat (UAE) 100% - 24 agents
    add_rollout(datetime(2026, 2, 9), datetime(2026, 2, 10), "Partner Care (L1)", "Chat", "UAE", 24)
    
    # 11-12 Feb: Phone (KSA) 100% - 26 agents
    add_rollout(datetime(2026, 2, 11), datetime(2026, 2, 12), "Partner Care (L1)", "Phone", "KSA", 26)
    
    # 11-12 Feb: Phone (UAE) 100% - 5 agents
    add_rollout(datetime(2026, 2, 11), datetime(2026, 2, 12), "Partner Care (L1)", "Phone", "UAE", 5)
    
    # 11-15 Feb: Chat (KSA) 50% - 8 agents
    add_rollout(datetime(2026, 2, 11), datetime(2026, 2, 15), "Partner Care (L1)", "Chat", "KSA", 8)
    
    # 16 Feb: Chat (KSA) 100% - 15 agents
    add_rollout(datetime(2026, 2, 16), datetime(2026, 2, 16), "Partner Care (L1)", "Chat", "KSA", 15)
    
    # ============================================================
    # PARTNER ONBOARDING (L1) - UPDATED (44 agents for email)
    # ============================================================
    
    # 02/02/2026: Email (UAE + KSA) - 44 agents total (split ~22/22)
    add_rollout(datetime(2026, 2, 2), datetime(2026, 2, 2), "Partner Onboarding (L1)", "Email", "UAE", 22)
    add_rollout(datetime(2026, 2, 2), datetime(2026, 2, 2), "Partner Onboarding (L1)", "Email", "KSA", 22)
    
    # 02/02/2026 → 04/02/2026: Soft launch - 2 agents each
    add_rollout(datetime(2026, 2, 2), datetime(2026, 2, 4), "Partner Onboarding (L1)", "Phone", "KSA", 2)
    add_rollout(datetime(2026, 2, 2), datetime(2026, 2, 4), "Partner Onboarding (L1)", "Chat", "KSA", 2)
    add_rollout(datetime(2026, 2, 2), datetime(2026, 2, 4), "Partner Onboarding (L1)", "Phone", "UAE", 2)
    add_rollout(datetime(2026, 2, 2), datetime(2026, 2, 4), "Partner Onboarding (L1)", "Chat", "UAE", 2)
    
    # 05-06 Feb: Chat (UAE) 20% - 3 agents
    add_rollout(datetime(2026, 2, 5), datetime(2026, 2, 6), "Partner Onboarding (L1)", "Chat", "UAE", 3)
    
    # 07-08 Feb: Chat (UAE) 50% - 7 agents
    add_rollout(datetime(2026, 2, 7), datetime(2026, 2, 8), "Partner Onboarding (L1)", "Chat", "UAE", 7)
    
    # 09-10 Feb: Chat (UAE) 100% - 12 agents
    add_rollout(datetime(2026, 2, 9), datetime(2026, 2, 10), "Partner Onboarding (L1)", "Chat", "UAE", 12)
    
    # 11-12 Feb: Phone (UAE+KSA) 100% - 2 agents total
    add_rollout(datetime(2026, 2, 11), datetime(2026, 2, 12), "Partner Onboarding (L1)", "Phone", "UAE", 1)
    add_rollout(datetime(2026, 2, 11), datetime(2026, 2, 12), "Partner Onboarding (L1)", "Phone", "KSA", 1)
    
    # 11-15 Feb: Chat (KSA) 50% - 19 agents
    add_rollout(datetime(2026, 2, 11), datetime(2026, 2, 15), "Partner Onboarding (L1)", "Chat", "KSA", 19)
    
    # 16 Feb: Chat (KSA) 100% - 33 agents
    add_rollout(datetime(2026, 2, 16), datetime(2026, 2, 16), "Partner Onboarding (L1)", "Chat", "KSA", 33)
    
    # ============================================================
    # CUSTOMER CARE (L1) (STANDARD) - UPDATED
    # ============================================================
    
    # 02/02/2026 → 04/02/2026: Soft launch
    add_rollout(datetime(2026, 2, 2), datetime(2026, 2, 4), "Customer Care (L1) (Standard)", "Phone", "KSA", 5)
    add_rollout(datetime(2026, 2, 2), datetime(2026, 2, 4), "Customer Care (L1) (Standard)", "Chat", "KSA", 5)
    add_rollout(datetime(2026, 2, 2), datetime(2026, 2, 4), "Customer Care (L1) (Standard)", "Email", "KSA", 2)
    add_rollout(datetime(2026, 2, 2), datetime(2026, 2, 4), "Customer Care (L1) (Standard)", "Phone", "UAE", 2)
    add_rollout(datetime(2026, 2, 2), datetime(2026, 2, 4), "Customer Care (L1) (Standard)", "Chat", "UAE", 2)
    add_rollout(datetime(2026, 2, 2), datetime(2026, 2, 4), "Customer Care (L1) (Standard)", "Email", "UAE", 2)
    
    # 09-12 Feb: Email (UAE + KSA) 100% - 17 agents (split 8/9)
    add_rollout(datetime(2026, 2, 9), datetime(2026, 2, 12), "Customer Care (L1) (Standard)", "Email", "UAE", 8)
    add_rollout(datetime(2026, 2, 9), datetime(2026, 2, 12), "Customer Care (L1) (Standard)", "Email", "KSA", 9)
    
    # 13-16 Feb: Chat (UAE) 20% - 3 agents
    add_rollout(datetime(2026, 2, 13), datetime(2026, 2, 16), "Customer Care (L1) (Standard)", "Chat", "UAE", 3)
    
    # 13-16 Feb: Phone (UAE) 100% - 10 agents
    add_rollout(datetime(2026, 2, 13), datetime(2026, 2, 16), "Customer Care (L1) (Standard)", "Phone", "UAE", 10)
    
    # 17-19 Feb: Chat (UAE) 50% - 7 agents
    add_rollout(datetime(2026, 2, 17), datetime(2026, 2, 19), "Customer Care (L1) (Standard)", "Chat", "UAE", 7)
    
    # 17-19 Feb: Chat (KSA) 20% - 20 agents
    add_rollout(datetime(2026, 2, 17), datetime(2026, 2, 19), "Customer Care (L1) (Standard)", "Chat", "KSA", 20)
    
    # 17-19 Feb: Phone (KSA) 50% - 41 agents (NEW: split into two phases)
    add_rollout(datetime(2026, 2, 17), datetime(2026, 2, 19), "Customer Care (L1) (Standard)", "Phone", "KSA", 41)
    
    # 20-22 Feb: Phone (KSA) 100% - 76 agents (NEW: second phase)
    add_rollout(datetime(2026, 2, 20), datetime(2026, 2, 22), "Customer Care (L1) (Standard)", "Phone", "KSA", 76)
    
    # 20-22 Feb: Chat (UAE) 100% - 12 agents
    add_rollout(datetime(2026, 2, 20), datetime(2026, 2, 22), "Customer Care (L1) (Standard)", "Chat", "UAE", 12)
    
    # 20-22 Feb: Chat (KSA) 50% - 45 agents
    add_rollout(datetime(2026, 2, 20), datetime(2026, 2, 22), "Customer Care (L1) (Standard)", "Chat", "KSA", 45)
    
    # 23-24 Feb: Chat (KSA) 100% - 83 agents
    add_rollout(datetime(2026, 2, 23), datetime(2026, 2, 24), "Customer Care (L1) (Standard)", "Chat", "KSA", 83)
    
    # ============================================================
    # CUSTOMER CARE (L1) (SMART)
    # ============================================================
    
    # 02/02/2026 → 04/02/2026: Soft launch
    add_rollout(datetime(2026, 2, 2), datetime(2026, 2, 4), "Customer Care (L1) (Smart)", "Phone", "UAE", 2)
    add_rollout(datetime(2026, 2, 2), datetime(2026, 2, 4), "Customer Care (L1) (Smart)", "Chat", "UAE", 2)
    add_rollout(datetime(2026, 2, 2), datetime(2026, 2, 4), "Customer Care (L1) (Smart)", "Email", "KSA", 2)
    
    # 09-12 Feb: Email (UAE + KSA) 100% - 2 agents
    add_rollout(datetime(2026, 2, 9), datetime(2026, 2, 12), "Customer Care (L1) (Smart)", "Email", "UAE", 1)
    add_rollout(datetime(2026, 2, 9), datetime(2026, 2, 12), "Customer Care (L1) (Smart)", "Email", "KSA", 1)
    
    # 13-16 Feb: Chat (UAE) 20% - 3 agents
    add_rollout(datetime(2026, 2, 13), datetime(2026, 2, 16), "Customer Care (L1) (Smart)", "Chat", "UAE", 3)
    
    # 13-16 Feb: Phone (UAE) 100% - 12 agents
    add_rollout(datetime(2026, 2, 13), datetime(2026, 2, 16), "Customer Care (L1) (Smart)", "Phone", "UAE", 12)
    
    # 17-19 Feb: Chat (UAE) 50% - 6 agents
    add_rollout(datetime(2026, 2, 17), datetime(2026, 2, 19), "Customer Care (L1) (Smart)", "Chat", "UAE", 6)
    
    # 20-22 Feb: Chat (UAE) 100% - 10 agents
    add_rollout(datetime(2026, 2, 20), datetime(2026, 2, 22), "Customer Care (L1) (Smart)", "Chat", "UAE", 10)
    
    # KSA Smart has 0 agents (no volume)
    
    # ============================================================
    # PARTNER CARE L2 TEAMS - NEW DATA
    # ============================================================
    
    # 03/02/2026 → 05/02/2026: 0% → 50%
    add_rollout(datetime(2026, 2, 3), datetime(2026, 2, 5), "Partner Technical Support L2", "All", "Cross-country", 1)
    add_rollout(datetime(2026, 2, 3), datetime(2026, 2, 5), "Partner Trust & Safety L2", "All", "Cross-country", 1)
    add_rollout(datetime(2026, 2, 3), datetime(2026, 2, 5), "Partner Finance L2", "All", "Cross-country", 2)
    add_rollout(datetime(2026, 2, 3), datetime(2026, 2, 5), "Partner Manager Escalation L2", "All", "Cross-country", 4)
    
    # 06/02/2026 → 08/02/2026: 50% → 100%
    add_rollout(datetime(2026, 2, 6), datetime(2026, 2, 8), "Partner Technical Support L2", "All", "Cross-country", 1)
    add_rollout(datetime(2026, 2, 6), datetime(2026, 2, 8), "Partner Trust & Safety L2", "All", "Cross-country", 2)
    add_rollout(datetime(2026, 2, 6), datetime(2026, 2, 8), "Partner Finance L2", "All", "Cross-country", 3)
    add_rollout(datetime(2026, 2, 6), datetime(2026, 2, 8), "Partner Manager Escalation L2", "All", "Cross-country", 4)
    
    # ============================================================
    # CUSTOMER L2 TEAMS - NEW DATA
    # ============================================================
    
    # 9/02/2026 → 12/02/2026: 0% → 20%
    add_rollout(datetime(2026, 2, 9), datetime(2026, 2, 12), "Customer Care L2 (SME)", "All", "Cross-country", 3)
    add_rollout(datetime(2026, 2, 9), datetime(2026, 2, 12), "Customer Govt & Exec Escalations L2", "All", "Cross-country", 2)
    add_rollout(datetime(2026, 2, 9), datetime(2026, 2, 12), "Customer Manager Escalations L2", "All", "Cross-country", 4)
    add_rollout(datetime(2026, 2, 9), datetime(2026, 2, 12), "Customer Technical Support L2", "All", "Cross-country", 2)
    add_rollout(datetime(2026, 2, 9), datetime(2026, 2, 12), "Customer Trust & Safety L2", "All", "Cross-country", 1)
    
    # 13/02/2026 → 16/02/2026: 20% → 70%
    add_rollout(datetime(2026, 2, 13), datetime(2026, 2, 16), "Customer Care L2 (SME)", "All", "Cross-country", 9)
    add_rollout(datetime(2026, 2, 13), datetime(2026, 2, 16), "Customer Govt & Exec Escalations L2", "All", "Cross-country", 6)
    add_rollout(datetime(2026, 2, 13), datetime(2026, 2, 16), "Customer Manager Escalations L2", "All", "Cross-country", 13)
    add_rollout(datetime(2026, 2, 13), datetime(2026, 2, 16), "Customer Technical Support L2", "All", "Cross-country", 6)
    add_rollout(datetime(2026, 2, 13), datetime(2026, 2, 16), "Customer Trust & Safety L2", "All", "Cross-country", 2)
    
    # 17/02/2026 → 19/02/2026: 70% → 100%
    add_rollout(datetime(2026, 2, 17), datetime(2026, 2, 19), "Customer Care L2 (SME)", "All", "Cross-country", 12)
    add_rollout(datetime(2026, 2, 17), datetime(2026, 2, 19), "Customer Govt & Exec Escalations L2", "All", "Cross-country", 9)
    add_rollout(datetime(2026, 2, 17), datetime(2026, 2, 19), "Customer Manager Escalations L2", "All", "Cross-country", 18)
    add_rollout(datetime(2026, 2, 17), datetime(2026, 2, 19), "Customer Technical Support L2", "All", "Cross-country", 9)
    add_rollout(datetime(2026, 2, 17), datetime(2026, 2, 19), "Customer Trust & Safety L2", "All", "Cross-country", 3)
    
    return rollout_data


def create_excel():
    """Create the Excel workbook with rollout schedule."""
    
    print("Creating rollout schedule Excel (UPDATED)...")
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SF Rollout Feb 2026"
    
    # Define styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    date_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Category colors
    partner_care_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
    partner_onb_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")   # Light orange
    customer_std_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # Light blue
    customer_smart_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Light mint
    l2_partner_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Light red
    l2_customer_fill = PatternFill(start_color="D9D2E9", end_color="D9D2E9", fill_type="solid")   # Light purple
    support_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")       # Light pink
    
    # Get rollout data
    rollout_data = create_rollout_data()
    
    # Create list of all LOB keys for columns
    lob_columns = []
    for lob, channel, country in LOBS:
        lob_key = get_lob_key(lob, channel, country)
        lob_columns.append((lob_key, lob, channel, country))
    
    # Header row 1: Categories
    ws.cell(row=1, column=1, value="Date")
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).alignment = center_align
    ws.cell(row=1, column=1).border = thin_border
    
    # Merge cells for category headers and add LOB headers
    col = 2
    category_ranges = []
    current_lob = None
    start_col = 2
    
    for i, (lob_key, lob, channel, country) in enumerate(lob_columns):
        if lob != current_lob:
            if current_lob is not None:
                category_ranges.append((current_lob, start_col, col - 1))
            current_lob = lob
            start_col = col
        col += 1
    # Don't forget the last category
    if current_lob is not None:
        category_ranges.append((current_lob, start_col, col - 1))
    
    # Write category headers (row 1) with merges
    for cat_name, start, end in category_ranges:
        if start < end:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        cell = ws.cell(row=1, column=start, value=cat_name)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
        # Set fill color based on category
        if "Partner Care (L1)" in cat_name:
            cell.fill = partner_care_fill
        elif "Partner Onboarding" in cat_name:
            cell.fill = partner_onb_fill
        elif "Customer Care (L1) (Standard)" in cat_name:
            cell.fill = customer_std_fill
        elif "Customer Care (L1) (Smart)" in cat_name:
            cell.fill = customer_smart_fill
        elif "Partner" in cat_name and "L2" in cat_name:
            cell.fill = l2_partner_fill
        elif "Customer" in cat_name and "L2" in cat_name:
            cell.fill = l2_customer_fill
        else:
            cell.fill = support_fill
    
    # Header row 2: Channel - Country
    ws.cell(row=2, column=1, value="")
    ws.cell(row=2, column=1).fill = header_fill
    ws.cell(row=2, column=1).border = thin_border
    
    col = 2
    for lob_key, lob, channel, country in lob_columns:
        cell = ws.cell(row=2, column=col, value=f"{channel}\n({country})")
        cell.font = Font(bold=True, size=9)
        cell.alignment = center_align
        cell.border = thin_border
        
        # Set fill color
        if "Partner Care (L1)" in lob:
            cell.fill = partner_care_fill
        elif "Partner Onboarding" in lob:
            cell.fill = partner_onb_fill
        elif "Customer Care (L1) (Standard)" in lob:
            cell.fill = customer_std_fill
        elif "Customer Care (L1) (Smart)" in lob:
            cell.fill = customer_smart_fill
        elif "Partner" in lob and "L2" in lob:
            cell.fill = l2_partner_fill
        elif "Customer" in lob and "L2" in lob:
            cell.fill = l2_customer_fill
        else:
            cell.fill = support_fill
        
        col += 1
    
    # Generate all February 2026 dates
    feb_dates = []
    current_date = datetime(2026, 2, 1)
    while current_date.month == 2:
        feb_dates.append(current_date)
        current_date += timedelta(days=1)
    
    # Fill in the data rows
    for row_idx, date in enumerate(feb_dates, start=3):
        # Date column
        date_cell = ws.cell(row=row_idx, column=1, value=date.strftime("%d/%m/%Y (%a)"))
        date_cell.fill = date_fill
        date_cell.alignment = Alignment(horizontal="left", vertical="center")
        date_cell.border = thin_border
        date_cell.font = Font(bold=True, size=10)
        
        # Highlight weekends
        if date.weekday() >= 5:  # Saturday or Sunday
            date_cell.fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        
        # LOB columns
        col = 2
        for lob_key, lob, channel, country in lob_columns:
            agent_count = rollout_data[date].get(lob_key, 0)
            cell = ws.cell(row=row_idx, column=col, value=agent_count if agent_count > 0 else "")
            cell.alignment = center_align
            cell.border = thin_border
            
            # Highlight cells with agents
            if agent_count > 0:
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                cell.font = Font(bold=True)
            
            col += 1
    
    # Add totals row
    total_row = len(feb_dates) + 3
    ws.cell(row=total_row, column=1, value="MAX AGENTS").font = Font(bold=True)
    ws.cell(row=total_row, column=1).fill = header_fill
    ws.cell(row=total_row, column=1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=total_row, column=1).border = thin_border
    
    col = 2
    for lob_key, lob, channel, country in lob_columns:
        # Find max agents for this LOB across all dates
        max_agents = max([rollout_data[date].get(lob_key, 0) for date in feb_dates], default=0)
        cell = ws.cell(row=total_row, column=col, value=max_agents if max_agents > 0 else "")
        cell.font = Font(bold=True)
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        col += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 18
    for col_idx in range(2, len(lob_columns) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12
    
    # Freeze panes (freeze date column and first two header rows)
    ws.freeze_panes = "B3"
    
    # Create summary sheet
    ws2 = wb.create_sheet(title="Summary by Date")
    
    # Summary headers
    summary_headers = ["Date", "Day", "Total Agents on SF", "Partner Care L1", "Partner Onboarding L1", 
                       "Customer Care Std L1", "Customer Care Smart L1", "Partner L2", "Customer L2", "Support"]
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # Summary data
    for row_idx, date in enumerate(feb_dates, start=2):
        # Calculate totals by category
        total = 0
        partner_care_total = 0
        partner_onb_total = 0
        customer_std_total = 0
        customer_smart_total = 0
        partner_l2_total = 0
        customer_l2_total = 0
        support_total = 0
        
        for lob_key, lob, channel, country in lob_columns:
            count = rollout_data[date].get(lob_key, 0)
            total += count
            
            if "Partner Care (L1)" in lob:
                partner_care_total += count
            elif "Partner Onboarding" in lob:
                partner_onb_total += count
            elif "Customer Care (L1) (Standard)" in lob:
                customer_std_total += count
            elif "Customer Care (L1) (Smart)" in lob:
                customer_smart_total += count
            elif "Partner" in lob and "L2" in lob:
                partner_l2_total += count
            elif "Customer" in lob and "L2" in lob:
                customer_l2_total += count
            else:
                support_total += count
        
        ws2.cell(row=row_idx, column=1, value=date.strftime("%d/%m/%Y")).border = thin_border
        ws2.cell(row=row_idx, column=2, value=date.strftime("%A")).border = thin_border
        ws2.cell(row=row_idx, column=3, value=total).border = thin_border
        ws2.cell(row=row_idx, column=4, value=partner_care_total).border = thin_border
        ws2.cell(row=row_idx, column=5, value=partner_onb_total).border = thin_border
        ws2.cell(row=row_idx, column=6, value=customer_std_total).border = thin_border
        ws2.cell(row=row_idx, column=7, value=customer_smart_total).border = thin_border
        ws2.cell(row=row_idx, column=8, value=partner_l2_total).border = thin_border
        ws2.cell(row=row_idx, column=9, value=customer_l2_total).border = thin_border
        ws2.cell(row=row_idx, column=10, value=support_total).border = thin_border
        
        # Highlight weekends
        if date.weekday() >= 5:
            for col_idx in range(1, 11):
                ws2.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    start_color="F4B084", end_color="F4B084", fill_type="solid"
                )
    
    # Adjust summary column widths
    for col_idx in range(1, 11):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 18
    
    ws2.freeze_panes = "A2"
    
    # Create third sheet with milestones - UPDATED
    ws3 = wb.create_sheet(title="Key Milestones")
    
    milestones = [
        ("02/02/2026", "Partner Onboarding - Email Launch (UAE + KSA)", "44 agents - Email cutover 100%"),
        ("02/02/2026", "All LOBs - Soft Launch (Phone/Chat/Email)", "2-5 agents per LOB for testing"),
        ("03/02/2026", "Partner Care - Email Launch (UAE + KSA)", "6 agents - Email cutover 100%"),
        ("03/02/2026", "Partner L2 Teams - 50% Rollout Start", "Tech Support=1, T&S=1, Finance=2, Mgr Esc=4"),
        ("05/02/2026", "Partner Care - Chat UAE 20%", "8 agents - Gradual rollout starts"),
        ("06/02/2026", "Partner L2 Teams - 100% Complete", "Tech Support=1, T&S=2, Finance=3, Mgr Esc=4"),
        ("09/02/2026", "Partner Care - Chat UAE 100%", "24 agents - UAE Chat complete"),
        ("09/02/2026", "Customer Care (Standard + Smart) - Email Launch", "Email cutover 100%"),
        ("09/02/2026", "Customer L2 Teams - 20% Rollout Start", "SME=3, Govt=2, Mgr Esc=4, Tech=2, T&S=1"),
        ("11/02/2026", "Partner Care - Phone Launch (UAE + KSA)", "UAE=5, KSA=26 agents"),
        ("13/02/2026", "Customer Care - Phone UAE Launch", "Standard=10, Smart=12 agents"),
        ("13/02/2026", "Customer L2 Teams - 70%", "SME=9, Govt=6, Mgr Esc=13, Tech=6, T&S=2"),
        ("16/02/2026", "Partner Care - Chat KSA 100%", "15 agents - KSA Chat complete"),
        ("17/02/2026", "Customer Care - Phone KSA 50%", "41 agents - First wave"),
        ("17/02/2026", "Customer L2 Teams - 100% Complete", "SME=12, Govt=9, Mgr Esc=18, Tech=9, T&S=3"),
        ("20/02/2026", "Customer Care - Phone KSA 100%", "76 agents - KSA Phone complete"),
        ("22/02/2026", "Customer Care - Chat UAE 100%", "Standard=12, Smart=10 agents"),
        ("24/02/2026", "Customer Care - Chat KSA 100%", "83 agents - FULL L1 ROLLOUT COMPLETE"),
    ]
    
    milestone_headers = ["Date", "Milestone", "Details"]
    for col_idx, header in enumerate(milestone_headers, start=1):
        cell = ws3.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    for row_idx, (date, milestone, notes) in enumerate(milestones, start=2):
        ws3.cell(row=row_idx, column=1, value=date).border = thin_border
        ws3.cell(row=row_idx, column=2, value=milestone).border = thin_border
        ws3.cell(row=row_idx, column=3, value=notes).border = thin_border
    
    ws3.column_dimensions['A'].width = 15
    ws3.column_dimensions['B'].width = 55
    ws3.column_dimensions['C'].width = 50
    
    # Save workbook
    wb.save(OUTPUT_FILE)
    print(f"\n✅ Excel file created successfully: {OUTPUT_FILE}")
    print(f"\nSheets created:")
    print(f"  1. SF Rollout Feb 2026 - Day-by-day agent counts per LOB")
    print(f"  2. Summary by Date - Daily totals by category")
    print(f"  3. Key Milestones - Important rollout dates (UPDATED)")
    
    # Print summary
    print(f"\n📊 Summary:")
    print(f"  - Total LOB columns: {len(lob_columns)}")
    print(f"  - Date range: Feb 1-28, 2026")
    
    max_day_total = max([sum(rollout_data[date].values()) for date in feb_dates])
    max_day = max(feb_dates, key=lambda d: sum(rollout_data[d].values()))
    print(f"  - Max agents on any single day: {max_day_total} (on {max_day.strftime('%d/%m/%Y')})")
    
    # Print L2 summary
    print(f"\n📋 L2 Teams Summary:")
    print(f"  Partner L2 (from 06 Feb): Tech Support=1, T&S=2, Finance=3, Mgr Esc=4 → Total: 10")
    print(f"  Customer L2 (from 17 Feb): SME=12, Govt=9, Mgr Esc=18, Tech=9, T&S=3 → Total: 51")
    
    return OUTPUT_FILE


if __name__ == "__main__":
    create_excel()
